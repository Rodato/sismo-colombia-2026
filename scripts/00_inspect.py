"""Inspecciona la estructura de cada fuente antes de cruzarlas.

No calcula nada: solo imprime columnas, tipos y rangos para poder escribir
el parseo real sin adivinar.
"""

import warnings
from pathlib import Path

import geopandas as gpd
import openpyxl

warnings.filterwarnings("ignore")

RAW = Path(__file__).resolve().parent.parent / "data" / "raw"


def sep(t):
    print(f"\n{'=' * 70}\n{t}\n{'=' * 70}")


# ---------------------------------------------------------------- MGN
sep("MGN 2023 — municipios DANE")
mgn = gpd.read_file(RAW / "mgn_mpio" / "MGN_ADM_MPIO_GRAFICO.shp")
print(f"registros: {len(mgn)}   crs: {mgn.crs}")
print(f"columnas: {list(mgn.columns)}")
print(mgn.drop(columns="geometry").head(3).to_string())
print(f"\nbounds: {mgn.total_bounds}")


# ---------------------------------------------------------------- xlsx
def peek(fname, nrows=14, ncols=12):
    sep(f"XLSX — {fname}")
    wb = openpyxl.load_workbook(RAW / fname, read_only=True, data_only=True)
    print(f"hojas: {wb.sheetnames}")
    for sh in wb.sheetnames[:3]:
        ws = wb[sh]
        print(f"\n--- hoja '{sh}'  ({ws.max_row} filas x {ws.max_column} cols) ---")
        for i, row in enumerate(ws.iter_rows(max_row=nrows, max_col=ncols, values_only=True)):
            cells = ["" if c is None else str(c)[:20] for c in row]
            print(f"  r{i:<3}| " + " | ".join(cells))
    wb.close()


peek("dane_poblacion_mun_2018_2042.xlsx")
peek("dane_deficit_hab_2018_cnpv.xlsx")
peek("dane_deficit_hab_2021.xlsx")
